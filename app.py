# tsukasamiyashita/checkdailyreports/CheckDailyReports-94e7e606e86357cba1e8adce22bf563302ea0859/app.py
import os
import re
import sys
import json
import threading
from datetime import datetime
import tkinter as tk
from tkinter import filedialog, messagebox, ttk
import openpyxl
from openpyxl.utils import column_index_from_string

def resource_path(relative_path):
    """ Get absolute path to resource, works for dev and for PyInstaller """
    try:
        # PyInstaller creates a temp folder and stores path in _MEIPASS
        base_path = sys._MEIPASS
    except Exception:
        # スクリプトファイルが置かれているディレクトリを基準とする
        base_path = os.path.dirname(os.path.abspath(__file__))

    return os.path.join(base_path, relative_path)

def get_app_data_dir():
    """ 実行ファイルまたはスクリプトのあるディレクトリ配下の「CheckDailyReportsSettings」フォルダパスを取得し、なければ作成する """
    if getattr(sys, 'frozen', False):
        # PyInstaller環境 (.exe実行時)
        base_dir = os.path.dirname(sys.executable)
    else:
        # 通常のスクリプト実行時
        base_dir = os.path.dirname(os.path.abspath(__file__))
    
    target_dir = os.path.join(base_dir, "CheckDailyReportsSettings")
    if not os.path.exists(target_dir):
        try:
            os.makedirs(target_dir)
        except Exception as e:
            print(f"Failed to create directory {target_dir}: {e}")
    return target_dir

class DailyReportCheckerApp:
    def __init__(self, root):
        self.root = root
        self.root.title("CheckDailyReports-v1.1.0")
        self.root.geometry("900x750")
        self.root.minsize(800, 600)
        
        # Windowsのタスクバーにカスタムアイコンを正しく表示するための処理
        try:
            import ctypes
            myappid = 'tsukasamiyashita.checkdailyreports.checker.v1.1.0'
            ctypes.windll.shell32.SetCurrentProcessExplicitAppUserModelID(myappid)
        except Exception:
            pass
            
        # アイコンの設定 (default引数を使用してすべてのToplevelウィンドウへ適用)
        icon_path = resource_path("icon.ico")
        if os.path.exists(icon_path):
            try:
                self.root.iconbitmap(default=icon_path)
            except Exception:
                pass
        
        # 起動時にウィンドウを最大化
        try:
            self.root.state('zoomed')
        except tk.TclError:
            # Linux等の環境でzoomedがサポートされていない場合のフォールバック
            self.root.attributes('-zoomed', True)

        # スタイル設定
        self.style = ttk.Style()
        self.style.theme_use("clam")
        
        # 配色定義
        self.bg_color = "#f3f4f6"
        self.primary_color = "#2563eb"
        self.accent_color = "#dc2626"
        
        self.root.configure(bg=self.bg_color)
        
        # Treeviewの表示スタイルカスタム（行の高さを26pxにし、文字フォントを調整）
        self.style.configure("Treeview", 
                             rowheight=26, 
                             font=("Yu Gothic UI", 10))
        self.style.configure("Treeview.Heading", 
                             font=("Yu Gothic UI", 10, "bold"))
        
        # 状態保持変数
        self.template_file = tk.StringVar() # 基準テンプレートファイルのパス
        
        # 行・列の設定変数（UI非表示でもデフォルト値として機能）
        self.start_row_var = tk.StringVar(value="8")
        self.end_row_var = tk.StringVar(value="158")
        self.mng_col_var = tk.StringVar(value="B")
        self.kat_col_var = tk.StringVar(value="I")
        self.sag_col_var = tk.StringVar(value="K")
        self.tot_col_var = tk.StringVar(value="AR")

        # チェック項目別有効/無効設定変数
        self.cfg_check_smart_scan = tk.BooleanVar(value=True)       # 氏名・日付等スマートスキャン
        self.cfg_check_mng_no = tk.BooleanVar(value=True)          # 管理No未入力チェック
        self.cfg_check_code_integrity = tk.BooleanVar(value=True)  # コード未入力チェック（型枠・作業コード）
        self.cfg_check_numeric_format = tk.BooleanVar(value=True)  # 数値フォーマットチェック
        self.cfg_check_time_multiples = tk.BooleanVar(value=True)  # 時間単位（設定された倍数）チェック
        self.cfg_check_template_compare = tk.BooleanVar(value=True)# 基準テンプレート不整合チェック
        
        # 勤務時間単位チェックの倍数(分)設定
        self.cfg_time_multiple_val = tk.StringVar(value="60")
        
        self.is_processing = False
        self.cancel_requested = False
        self.template_cache = {} # テンプレートセルのキャッシュ
        
        # フィルター用データ保持
        self.all_error_data = [] # 検出された全エラーデータをメモリ保持
        self.detected_error_types = set(["すべて"]) # 検出されたユニークなエラー区分の保持用

        # UIの構築
        self._build_ui()

        # メニューバーの構築
        self._build_menu()

        # 保存されているデフォルト設定の読み込み
        self._load_config()

        # ウィンドウを閉じるイベントを設定
        self.root.protocol("WM_DELETE_WINDOW", self._on_closing)

    def _build_ui(self):
        # メインコンテナ
        main_frame = ttk.Frame(self.root, padding=15)
        main_frame.pack(fill=tk.BOTH, expand=True)

        # ヘッダーエリア
        header_frame = ttk.Frame(main_frame)
        header_frame.pack(fill=tk.X, pady=(0, 10))
        
        title_label = tk.Label(
            header_frame, 
            text="日報入力ミス自動チェックツール", 
            font=("Yu Gothic UI", 16, "bold"),
            bg=self.bg_color,
            fg="#1f2937"
        )
        title_label.pack(side=tk.LEFT)
        
        ver_label = tk.Label(
            header_frame,
            text="v1.1.0",
            font=("Yu Gothic UI", 10, "italic"),
            bg=self.bg_color,
            fg="#6b7280"
        )
        ver_label.pack(side=tk.LEFT, padx=10, pady=(5, 0))

        # フォルダ選択エリア (複数フォルダ用リストボックスに変更)
        folder_frame = ttk.LabelFrame(main_frame, text=" 1. 対象フォルダ選択 (複数登録可能) ", padding=10)
        folder_frame.pack(fill=tk.X, pady=(0, 10))

        # リストボックスとスクロールバーのコンテナ
        list_container = ttk.Frame(folder_frame)
        list_container.pack(side=tk.LEFT, fill=tk.BOTH, expand=True, padx=(0, 5))

        self.folder_scrollbar = ttk.Scrollbar(list_container, orient=tk.VERTICAL)
        self.folder_scrollbar.pack(side=tk.RIGHT, fill=tk.Y)

        self.folder_listbox = tk.Listbox(
            list_container, 
            height=4, 
            font=("Yu Gothic UI", 10), 
            yscrollcommand=self.folder_scrollbar.set,
            selectmode=tk.SINGLE
        )
        self.folder_listbox.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        self.folder_scrollbar.config(command=self.folder_listbox.yview)

        # 操作ボタンのコンテナ (右側配置)
        btn_container = ttk.Frame(folder_frame)
        btn_container.pack(side=tk.RIGHT, fill=tk.Y)

        add_btn = ttk.Button(btn_container, text="追加 (＋)", command=self._add_folder)
        add_btn.pack(fill=tk.X, pady=2)

        remove_btn = ttk.Button(btn_container, text="削除 (ー)", command=self._remove_folder)
        remove_btn.pack(fill=tk.X, pady=2)

        save_folder_btn = ttk.Button(btn_container, text="保存", command=self._save_config_with_message)
        save_folder_btn.pack(fill=tk.X, pady=2)

        # 基準テンプレートファイル選択エリア
        template_frame = ttk.LabelFrame(main_frame, text=" 2. 基準テンプレートファイル選択 (入力以外の書き換えチェック用) ", padding=10)
        template_frame.pack(fill=tk.X, pady=(0, 10))

        self.template_entry = ttk.Entry(template_frame, textvariable=self.template_file, font=("Yu Gothic UI", 10))
        self.template_entry.pack(side=tk.LEFT, fill=tk.X, expand=True, padx=(0, 5))

        browse_temp_btn = ttk.Button(template_frame, text="参照...", command=self._browse_template)
        browse_temp_btn.pack(side=tk.LEFT, padx=(0, 5))

        save_temp_btn = ttk.Button(template_frame, text="保存", command=self._save_config_with_message)
        save_temp_btn.pack(side=tk.RIGHT)

        # 動作ボタンエリア
        btn_frame = ttk.Frame(main_frame)
        btn_frame.pack(fill=tk.X, pady=(0, 10))

        self.start_btn = tk.Button(
            btn_frame, 
            text="チェック開始", 
            bg=self.primary_color, 
            fg="white", 
            font=("Yu Gothic UI", 11, "bold"),
            activebackground="#1d4ed8",
            activeforeground="white",
            relief=tk.FLAT,
            padx=15,
            pady=6,
            command=self._start_check_thread
        )
        self.start_btn.pack(side=tk.LEFT, padx=(0, 10))

        self.stop_btn = tk.Button(
            btn_frame, 
            text="中止", 
            bg="#9ca3af", 
            fg="white", 
            font=("Yu Gothic UI", 11, "bold"),
            activebackground="#78716c",
            activeforeground="white",
            relief=tk.FLAT,
            state=tk.DISABLED,
            padx=15,
            pady=6,
            command=self._stop_check
        )
        self.stop_btn.pack(side=tk.LEFT, padx=(0, 10))

        # 設定画面を呼び出すボタンを追加
        self.settings_btn = tk.Button(
            btn_frame, 
            text="チェック項目設定...", 
            bg="#10b981", 
            fg="white", 
            font=("Yu Gothic UI", 10, "bold"),
            activebackground="#059669",
            activeforeground="white",
            relief=tk.FLAT,
            padx=15,
            pady=6,
            command=self._show_settings_popup
        )
        self.settings_btn.pack(side=tk.LEFT)

        # 進捗バーと進捗テキスト
        self.progress_frame = ttk.Frame(main_frame)
        self.progress_frame.pack(fill=tk.X, pady=(0, 10))
        
        self.progress_bar = ttk.Progressbar(self.progress_frame, orient="horizontal", mode="determinate")
        self.progress_bar.pack(fill=tk.X, side=tk.TOP, pady=(0, 2))
        
        self.status_label = tk.Label(
            self.progress_frame, 
            text="フォルダを選択して「チェック開始」をクリックしてください。", 
            font=("Yu Gothic UI", 9),
            bg=self.bg_color,
            fg="#4b5563",
            anchor="w"
        )
        self.status_label.pack(fill=tk.X, side=tk.BOTTOM)

        # 結果表示エリア（リスト）
        list_frame = ttk.LabelFrame(main_frame, text=" 3. チェック結果一覧 (エラー検出箇所) ", padding=10)
        list_frame.pack(fill=tk.BOTH, expand=True)

        # Excelライクなフィルターバーエリア
        filter_bar = ttk.Frame(list_frame, padding=(0, 0, 0, 10))
        filter_bar.pack(fill=tk.X, side=tk.TOP)

        ttk.Label(filter_bar, text="エラー区分で抽出: ", font=("Yu Gothic UI", 10)).pack(side=tk.LEFT, padx=(0, 5))
        self.filter_type_var = tk.StringVar(value="すべて")
        self.filter_type_combobox = ttk.Combobox(filter_bar, textvariable=self.filter_type_var, state="readonly", width=22, font=("Yu Gothic UI", 10))
        self.filter_type_combobox.pack(side=tk.LEFT, padx=(0, 20))
        self.filter_type_combobox.bind("<<ComboboxSelected>>", self._on_filter_changed)
        self.filter_type_combobox.config(values=["すべて"])

        ttk.Label(filter_bar, text="キーワード検索: ", font=("Yu Gothic UI", 10)).pack(side=tk.LEFT, padx=(0, 5))
        self.filter_keyword_var = tk.StringVar()
        self.filter_keyword_entry = ttk.Entry(filter_bar, textvariable=self.filter_keyword_var, width=25, font=("Yu Gothic UI", 10))
        self.filter_keyword_entry.pack(side=tk.LEFT, padx=(0, 8))
        self.filter_keyword_entry.bind("<KeyRelease>", self._on_filter_changed)

        clear_filter_btn = ttk.Button(filter_bar, text="フィルター解除", command=self._clear_filter)
        clear_filter_btn.pack(side=tk.LEFT)

        # スクロールバー
        scroll_y = ttk.Scrollbar(list_frame, orient=tk.VERTICAL)
        scroll_x = ttk.Scrollbar(list_frame, orient=tk.HORIZONTAL)

        # Treeview（カラム設定）
        cols = ("file_path", "sheet_name", "cell_pos", "error_type", "detail")
        self.tree = ttk.Treeview(
            list_frame, 
            columns=cols, 
            show="headings", 
            yscrollcommand=scroll_y.set, 
            xscrollcommand=scroll_x.set
        )
        
        # 縞模様用のタグ設定 (even: 白, odd: 少し濃く調整したソフトなグレー #e5e7eb)
        self.tree.tag_configure("even", background="#ffffff")
        self.tree.tag_configure("odd", background="#e5e7eb")
        
        scroll_y.config(command=self.tree.yview)
        scroll_y.pack(side=tk.RIGHT, fill=tk.Y)
        scroll_x.config(command=self.tree.xview)
        scroll_x.pack(side=tk.BOTTOM, fill=tk.X)
        self.tree.pack(fill=tk.BOTH, expand=True)

        # カラムヘッダー設定
        self.tree.heading("file_path", text="ファイル名 / パス")
        self.tree.heading("sheet_name", text="シート名")
        self.tree.heading("cell_pos", text="セル位置")
        self.tree.heading("error_type", text="エラー区分")
        self.tree.heading("detail", text="エラー詳細内容")

        # カラム幅調整
        self.tree.column("file_path", width=200, anchor="w")
        self.tree.column("sheet_name", width=100, anchor="center")
        self.tree.column("cell_pos", width=80, anchor="center")
        self.tree.column("error_type", width=130, anchor="center")
        self.tree.column("detail", width=400, anchor="w")

        # ダブルクリックイベント（詳細表示）
        self.tree.bind("<Double-1>", self._show_detail_popup)

    def _build_menu(self):
        """ メニューバーを構成する """
        self.menu_bar = tk.Menu(self.root)
        self.root.config(menu=self.menu_bar)

        # 設定メニューの追加
        self.settings_menu = tk.Menu(self.menu_bar, tearoff=0)
        self.menu_bar.add_cascade(label="設定", menu=self.settings_menu)
        self.settings_menu.add_command(label="チェック項目の設定", command=self._show_settings_popup)

        # ヘルプメニューの追加
        self.help_menu = tk.Menu(self.menu_bar, tearoff=0)
        self.menu_bar.add_cascade(label="ヘルプ", menu=self.help_menu)
        self.help_menu.add_command(label="ヘルプを表示 (README)", command=self._show_readme_help)

    def _show_settings_popup(self):
        """ チェック項目の有効化・無効化を設定するサブウィンドウを起動する """
        popup = tk.Toplevel(self.root)
        popup.title("チェック項目設定")
        popup.geometry("500x440")
        popup.resizable(False, False)
        popup.transient(self.root)
        popup.grab_set()
        popup.configure(bg="#f3f4f6")

        # タイトルラベル
        title_lbl = tk.Label(
            popup,
            text="🛠️ チェック項目の選択",
            font=("Yu Gothic UI", 12, "bold"),
            bg="#f3f4f6",
            fg="#1f2937",
            anchor="w"
        )
        title_lbl.pack(fill=tk.X, padx=20, pady=(15, 5))

        desc_lbl = tk.Label(
            popup,
            text="自動チェックを行う検証項目を選択してください。不要なチェックをオフにすることで、処理の高速化や特定のエラー検出を省略できます。",
            font=("Yu Gothic UI", 9),
            bg="#f3f4f6",
            fg="#4b5563",
            justify=tk.LEFT,
            wraplength=460,
            anchor="w"
        )
        desc_lbl.pack(fill=tk.X, padx=20, pady=(0, 15))

        # 設定オプション配置用のフレーム
        opts_frame = ttk.LabelFrame(popup, text=" 検証項目リスト ", padding=15)
        opts_frame.pack(fill=tk.BOTH, expand=True, padx=20, pady=5)

        # 各種チェックボックス (エラー区分名に整合させている)
        chk1 = ttk.Checkbutton(opts_frame, text="氏名・日付等スマートスキャンチェック (氏名/日付の記入漏れ、日付形式など)", variable=self.cfg_check_smart_scan)
        chk1.pack(anchor="w", pady=4)

        chk2 = ttk.Checkbutton(opts_frame, text="管理No未入力チェック (作業実績があるが管理Noが無い場合)", variable=self.cfg_check_mng_no)
        chk2.pack(anchor="w", pady=4)

        chk3 = ttk.Checkbutton(opts_frame, text="コード未入力チェック (管理No設定時、型枠・作業コードが未入力の場合)", variable=self.cfg_check_code_integrity)
        chk3.pack(anchor="w", pady=4)

        chk4 = ttk.Checkbutton(opts_frame, text="数値フォーマットチェック (必須入力エリア内の空白、数値以外を検出)", variable=self.cfg_check_numeric_format)
        chk4.pack(anchor="w", pady=4)

        # 時間単位（倍数指定対応）用のフレーム
        time_frame = ttk.Frame(opts_frame)
        time_frame.pack(anchor="w", pady=4, fill=tk.X)
        
        chk5 = ttk.Checkbutton(time_frame, text="勤務時間単位チェック (端数時間の入力エラーを検出)  倍数(分):", variable=self.cfg_check_time_multiples)
        chk5.pack(side=tk.LEFT)
        
        time_entry = ttk.Entry(time_frame, textvariable=self.cfg_time_multiple_val, width=5, font=("Yu Gothic UI", 10))
        time_entry.pack(side=tk.LEFT, padx=5)

        chk6 = ttk.Checkbutton(opts_frame, text="基準テンプレート不整合チェック (固定文字列や数式セルの書き換えを検出)", variable=self.cfg_check_template_compare)
        chk6.pack(anchor="w", pady=4)

        # フッターボタン
        btn_frame = ttk.Frame(popup, padding=(0, 15, 0, 15))
        btn_frame.pack(side=tk.BOTTOM, fill=tk.X)

        def save_and_close():
            # 倍数設定値のバリデーション
            val = self.cfg_time_multiple_val.get().strip()
            try:
                val_int = int(val)
                if val_int <= 0:
                    raise ValueError
            except ValueError:
                messagebox.showerror("エラー", "勤務時間単位チェックの倍数は、1以上の整数で入力してください。")
                return

            self._save_config()
            messagebox.showinfo("保存完了", "チェック項目の設定を「CheckDailyReportsSettings」フォルダ内に保存しました。")
            popup.destroy()

        save_btn = ttk.Button(btn_frame, text="設定を保存して閉じる", width=22, command=save_and_close)
        save_btn.pack(side=tk.LEFT, expand=True, padx=(20, 5))

        cancel_btn = ttk.Button(btn_frame, text="キャンセル", width=12, command=popup.destroy)
        cancel_btn.pack(side=tk.LEFT, expand=True, padx=(5, 20))

    def _show_readme_help(self):
        """ README.md の内容を美しく整形して表示するサブウィンドウを起動する """
        popup = tk.Toplevel(self.root)
        popup.title("ヘルプ (README)")
        popup.geometry("800x650")
        popup.minsize(600, 450)
        popup.transient(self.root)
        popup.grab_set()
        popup.configure(bg="#f3f4f6") # メイン背景色に合わせる

        # 上部タイトルヘッダー
        header_lbl = tk.Label(
            popup, 
            text="📖 ご利用ガイド・ヘルプ", 
            font=("Yu Gothic UI", 14, "bold"),
            bg="#f3f4f6",
            fg="#1f2937",
            anchor="w"
        )
        header_lbl.pack(fill=tk.X, padx=20, pady=(15, 5))

        frm = ttk.Frame(popup, padding=(20, 5, 20, 10))
        frm.pack(fill=tk.BOTH, expand=True)

        # テキストボックスとスクロールバー
        scroll_y = ttk.Scrollbar(frm, orient=tk.VERTICAL)
        text_area = tk.Text(
            frm, 
            wrap=tk.WORD, 
            yscrollcommand=scroll_y.set,
            bg="#ffffff",
            bd=0,
            highlightthickness=1,
            highlightbackground="#e5e7eb",
            highlightcolor="#e5e7eb",
            padx=25,
            pady=25
        )
        scroll_y.config(command=text_area.yview)

        scroll_y.pack(side=tk.RIGHT, fill=tk.Y)
        text_area.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)

        # Windows環境で綺麗に表示されるフォントを指定
        font_family = "Yu Gothic UI"

        # スタイルの詳細設定
        text_area.tag_configure("title", font=(font_family, 16, "bold"), foreground="#1e3a8a", spacing1=10, spacing3=15)
        text_area.tag_configure("h1", font=(font_family, 13, "bold"), foreground="#2563eb", spacing1=20, spacing3=10)
        text_area.tag_configure("h2", font=(font_family, 11, "bold"), foreground="#1f2937", spacing1=15, spacing3=8)
        text_area.tag_configure("body", font=(font_family, 10), foreground="#374151", spacing3=6, spacing1=2)
        text_area.tag_configure("bold", font=(font_family, 10, "bold"), foreground="#111827")
        text_area.tag_configure("inline_code", font=("Consolas", 10), background="#f3f4f6", foreground="#dc2626")
        text_area.tag_configure("bullet", font=(font_family, 10), foreground="#374151", lmargin1=24, lmargin2=38, spacing3=6, spacing1=2)
        text_area.tag_configure("code", font=("Consolas", 10), background="#f3f4f6", foreground="#1f2937", lmargin1=20, lmargin2=20, spacing1=6, spacing3=6)
        text_area.tag_configure("quote", font=(font_family, 10, "italic"), foreground="#4b5563", background="#f9fafb", lmargin1=24, lmargin2=24, spacing3=6)
        text_area.tag_configure("hr", font=(font_family, 10), foreground="#d1d5db", spacing1=15, spacing3=15)

        # インライン装飾の優先度を上げる（ベースのbody等に打ち勝つため）
        text_area.tag_raise("bold")
        text_area.tag_raise("inline_code")

        # README.mdの検索と読み込み
        readme_path = resource_path("README.md")
        content = ""
        if os.path.exists(readme_path):
            try:
                with open(readme_path, "r", encoding="utf-8") as f:
                    content = f.read()
            except Exception as e:
                content = f"README.md の読み込みに失敗しました。\n\nエラー詳細:\n{str(e)}"
        else:
            fallback_dirs = [
                os.path.dirname(os.path.abspath(__file__)),
                os.getcwd()
            ]
            found = False
            for d in fallback_dirs:
                p = os.path.join(d, "README.md")
                if os.path.exists(p):
                    try:
                        with open(p, "r", encoding="utf-8") as f:
                            content = f.read()
                        found = True
                        break
                    except:
                        pass
            if not found:
                content = "# エラー\n\nREADME.md ファイルが見つかりませんでした。\nアプリケーションパッケージに README.md が正しく同封されているか確認してください。"

        # インラインマークダウンパーサー関数
        def _insert_formatted_inline(text_widget, raw_text, base_tag):
            parts = re.split(r'(\*\*.*?\*\*|`.*?`)', raw_text)
            for part in parts:
                if part.startswith('**') and part.endswith('**'):
                    text_widget.insert(tk.END, part[2:-2], ("bold", base_tag))
                elif part.startswith('`') and part.endswith('`'):
                    text_widget.insert(tk.END, f" {part[1:-1]} ", ("inline_code", base_tag))
                else:
                    text_widget.insert(tk.END, part, base_tag)

        # 行単位での簡易Markdownパーサー
        lines = content.split('\n')
        in_code_block = False
        
        for line in lines:
            stripped = line.strip()
            
            # コードブロック処理
            if stripped.startswith('```'):
                in_code_block = not in_code_block
                continue
                
            if in_code_block:
                text_area.insert(tk.END, line + '\n', "code")
                continue
                
            # 各要素のパース
            if stripped.startswith('# '):
                text_area.insert(tk.END, stripped[2:] + '\n', "title")
            elif stripped.startswith('## '):
                text_area.insert(tk.END, stripped[3:] + '\n', "h1")
            elif stripped.startswith('### '):
                text_area.insert(tk.END, stripped[4:] + '\n', "h2")
            elif stripped == '---':
                text_area.insert(tk.END, "―" * 55 + "\n", "hr")
            elif stripped.startswith('- ') or stripped.startswith('* '):
                bullet_text = line.split('-', 1)[1].strip() if stripped.startswith('- ') else line.split('*', 1)[1].strip()
                text_area.insert(tk.END, "•  ", "bullet")
                _insert_formatted_inline(text_area, bullet_text, "bullet")
                text_area.insert(tk.END, "\n")
            elif stripped.startswith('> '):
                quote_text = stripped[2:]
                text_area.insert(tk.END, "▎ ", "bold")
                _insert_formatted_inline(text_area, quote_text, "quote")
                text_area.insert(tk.END, "\n")
            else:
                if stripped == "":
                    text_area.insert(tk.END, "\n", "body")
                else:
                    _insert_formatted_inline(text_area, line, "body")
                    text_area.insert(tk.END, "\n")

        text_area.config(state=tk.DISABLED) # 読込専用に設定

        # ボタンコンテナ (余白をしっかりと確保したモダンな配置)
        btn_frame = ttk.Frame(popup, padding=(0, 0, 0, 15))
        btn_frame.pack(side=tk.BOTTOM, fill=tk.X)
        
        close_btn = ttk.Button(btn_frame, text="閉じる", width=15, command=popup.destroy)
        close_btn.pack()

    def _col_to_index(self, col_str):
        try:
            return column_index_from_string(col_str.upper().strip()) - 1
        except:
            return -1

    def _load_config(self):
        """ 設定ファイル(config.json)から保存されたパス設定、及びチェック項目の有効設定を読み込む """
        try:
            config_dir = get_app_data_dir()
            config_path = os.path.join(config_dir, "config.json")
            if os.path.exists(config_path):
                with open(config_path, "r", encoding="utf-8") as f:
                    data = json.load(f)
                    
                    # 複数フォルダリストの復元
                    target_dir_data = data.get("target_dir", "")
                    self.folder_listbox.delete(0, tk.END)
                    
                    if isinstance(target_dir_data, list):
                        target_dirs = target_dir_data
                    elif isinstance(target_dir_data, str):
                        target_dirs = [p.strip() for p in target_dir_data.split(";") if p.strip()]
                    else:
                        target_dirs = []
                        
                    for d in target_dirs:
                        if d and os.path.exists(d):
                            self.folder_listbox.insert(tk.END, d)
                            
                    self.template_file.set(data.get("template_file", ""))

                    # チェック有効/無効設定の復元（互換性のためにデフォルトTrue）
                    self.cfg_check_smart_scan.set(data.get("check_smart_scan", True))
                    self.cfg_check_mng_no.set(data.get("check_mng_no", True))
                    self.cfg_check_code_integrity.set(data.get("check_code_integrity", True))
                    self.cfg_check_numeric_format.set(data.get("check_numeric_format", True))
                    self.cfg_check_time_multiples.set(data.get("check_time_multiples", True))
                    self.cfg_check_template_compare.set(data.get("check_template_compare", True))
                    
                    # 勤務時間単位の倍数をロード（無い場合はデフォルト "60"）
                    self.cfg_time_multiple_val.set(data.get("time_multiple_val", "60"))
        except Exception as e:
            print(f"Failed to load config: {e}")

    def _save_config(self):
        """ 設定ファイル(config.json)へ現在のパス設定、及びチェック項目の有効設定を保存する """
        try:
            config_dir = get_app_data_dir()
            config_path = os.path.join(config_dir, "config.json")
            
            # リストボックス内の全てのフォルダパスを取得
            target_dirs = list(self.folder_listbox.get(0, tk.END))
            
            data = {
                "target_dir": target_dirs,
                "template_file": self.template_file.get(),
                "check_smart_scan": self.cfg_check_smart_scan.get(),
                "check_mng_no": self.cfg_check_mng_no.get(),
                "check_code_integrity": self.cfg_check_code_integrity.get(),
                "check_numeric_format": self.cfg_check_numeric_format.get(),
                "check_time_multiples": self.cfg_check_time_multiples.get(),
                "check_template_compare": self.cfg_check_template_compare.get(),
                "time_multiple_val": self.cfg_time_multiple_val.get()
            }
            with open(config_path, "w", encoding="utf-8") as f:
                json.dump(data, f, ensure_ascii=False, indent=4)
        except Exception as e:
            print(f"Failed to save config: {e}")

    def _save_config_with_message(self):
        """ 設定ファイルへ現在のパス設定を保存し、確認メッセージを表示する """
        self._save_config()
        messagebox.showinfo("保存完了", "参照先の設定を「CheckDailyReportsSettings」フォルダ内に保存しました。")

    def _on_closing(self):
        """ アプリ終了時の処理 """
        self._save_config()
        self.root.destroy()

    def _add_folder(self):
        """ 対象フォルダをリストボックスに追加する """
        selected = filedialog.askdirectory()
        if selected:
            path = os.path.abspath(selected)
            existing = self.folder_listbox.get(0, tk.END)
            if path not in existing:
                self.folder_listbox.insert(tk.END, path)

    def _remove_folder(self):
        """ リストボックスで選択されているフォルダを削除する """
        selected_indices = self.folder_listbox.curselection()
        if not selected_indices:
            messagebox.showwarning("警告", "削除するフォルダをリストから選択してください。")
            return
        for index in reversed(selected_indices):
            self.folder_listbox.delete(index)

    def _browse_template(self):
        selected = filedialog.askopenfilename(
            filetypes=[("Excel Files", "*.xlsx;*.xlsm"), ("All Files", "*.*")]
        )
        if selected:
            self.template_file.set(os.path.abspath(selected))

    def _stop_check(self):
        if self.is_processing:
            self.cancel_requested = True
            self.status_label.config(text="中止処理中...")
            self.stop_btn.config(state=tk.DISABLED)

    def _start_check_thread(self):
        targets = list(self.folder_listbox.get(0, tk.END))
        if not targets:
            messagebox.showwarning("警告", "対象フォルダが登録されていません。")
            return
            
        # 複数フォルダそれぞれの存在チェック
        invalid_targets = [t for t in targets if not os.path.exists(t)]
        if invalid_targets:
            messagebox.showerror("エラー", "以下のフォルダが存在しません:\n" + "\n".join(invalid_targets))
            return
            
        try:
            start_row_val = int(self.start_row_var.get().strip())
            if start_row_val < 1:
                raise ValueError
        except ValueError:
            messagebox.showerror("エラー", "チェック開始行は1以上の半角数字で入力してください。")
            return

        end_row_str = self.end_row_var.get().strip()
        end_row_val = None
        if end_row_str:
            try:
                end_row_val = int(end_row_str)
                if end_row_val < start_row_val:
                    messagebox.showerror("エラー", "チェック最終行は開始行以上の数字を入力してください。")
                    return
            except ValueError:
                messagebox.showerror("エラー", "チェック最終行は半角数字で入力してください。")
                return
                
        cols_info = {
            'mng': self._col_to_index(self.mng_col_var.get()),
            'kat': self._col_to_index(self.kat_col_var.get()),
            'sag': self._col_to_index(self.sag_col_var.get()),
            'tot': self._col_to_index(self.tot_col_var.get())
        }

        # 開始前に現在の設定パスを保存
        self._save_config()

        # テンプレートファイルの確認と読み込み (テンプレート書き換えチェックが有効な場合のみ)
        temp_file_path = self.template_file.get().strip()
        if temp_file_path and self.cfg_check_template_compare.get():
            if not os.path.exists(temp_file_path):
                messagebox.showerror("エラー", "指定された基準テンプレートファイルが存在しません。")
                return
            self.status_label.config(text="基準テンプレートファイルを解析中...")
            if not self._load_template_data(temp_file_path):
                messagebox.showerror("エラー", "基準テンプレートファイルの読み込みに失敗しました。xlsx形式であることを確認してください。")
                return
        else:
            self.template_cache = {} # クリア

        self.is_processing = True
        self.cancel_requested = False
        self.start_btn.config(state=tk.DISABLED, bg="#9ca3af")
        self.stop_btn.config(state=tk.NORMAL, bg=self.accent_color)
        self.settings_btn.config(state=tk.DISABLED)
        
        # リストとフィルターの初期化
        for item in self.tree.get_children():
            self.tree.delete(item)
        self.all_error_data = []
        self.detected_error_types = set(["すべて"])
        self.filter_type_combobox.config(values=["すべて"])
        self.filter_type_var.set("すべて")
        self.filter_keyword_var.set("")

        # スレッド起動 (展開されたリストを渡す)
        thread = threading.Thread(target=self._run_checker, args=(targets, start_row_val, end_row_val, cols_info), daemon=True)
        thread.start()

    def _load_template_data(self, temp_file_path):
        """基準テンプレートExcelファイルの全内容（数式は数式文字列のまま）をメモリに読み込んで保持する"""
        self.template_cache = {}
        ext = os.path.splitext(temp_file_path)[1].lower()
        try:
            if ext in (".xlsx", ".xlsm"):
                # xlsx / xlsm は data_only=False にして数式文字列（=LOOKUP(...)など）を直接テンプレートとしてロードする
                wb = openpyxl.load_workbook(temp_file_path, read_only=True, data_only=False)
                for sheet_name in wb.sheetnames:
                    if not self._should_check_sheet(sheet_name):
                        continue
                    sheet = wb[sheet_name]
                    self.template_cache[sheet_name] = []
                    for row in sheet.iter_rows(values_only=True):
                        row_data = [self._get_clean_value(cell) for cell in row]
                        self.template_cache[sheet_name].append(row_data)
                wb.close()
                return True
            else:
                return False
        except Exception as e:
            print(f"Template load error: {e}")
            return False

    def _run_checker(self, targets, start_row_val, end_row_val, cols_info):
        valid_extensions = (".xlsx", ".xlsm")
        files_to_check = []
        
        # 登録されたすべてのフォルダを探索
        for t in targets:
            if os.path.exists(t):
                for root_path, _, files in os.walk(t):
                    for file in files:
                        if file.lower().endswith(valid_extensions) and not file.startswith("~$"):
                            files_to_check.append(os.path.join(root_path, file))

        # ファイル重複を排除
        files_to_check = list(dict.fromkeys(files_to_check))

        total_files = len(files_to_check)
        if total_files == 0:
            self.root.after(0, self._finish_checker, 0, "対象のExcelファイル(xlsx/xlsm)が見つかりませんでした。")
            return

        errors_found = 0
        for i, filepath in enumerate(files_to_check):
            if self.cancel_requested:
                break

            # 進捗更新表示用の相対パス (一致する親フォルダの相対パスを取得)
            relative_name = os.path.basename(filepath)
            for t in targets:
                if filepath.startswith(t):
                    relative_name = os.path.relpath(filepath, t)
                    break

            progress_percent = int(((i + 1) / total_files) * 100)
            self.root.after(
                0, 
                self._update_progress, 
                progress_percent, 
                f"検証中 ({i+1}/{total_files}): {relative_name}"
            )

            # チェック実施
            errors = self._check_excel_file(filepath, start_row_val, end_row_val, cols_info)
            if errors:
                errors_found += len(errors)
                self.root.after(0, self._add_errors_to_list, filepath, errors)

        msg = "チェックを中止しました。" if self.cancel_requested else f"チェック完了。全 {total_files} ファイル中 {errors_found} 件のエラーを検出しました。"
        self.root.after(0, self._finish_checker, errors_found, msg)

    def _update_progress(self, percent, text):
        self.progress_bar["value"] = percent
        self.status_label.config(text=text)

    def _add_errors_to_list(self, filepath, errors):
        for err in errors:
            basename = os.path.basename(filepath)
            sheet = err.get("sheet", "不明")
            cell = err.get("cell", "N/A")
            err_type = err.get("type", "警告")
            detail = err.get("detail", "")

            error_item = {
                "filepath": filepath,
                "basename": basename,
                "sheet": sheet,
                "cell": cell,
                "type": err_type,
                "detail": detail
            }
            # メモリ内に保存
            self.all_error_data.append(error_item)

            # エラー区分のコンボボックス選択肢を動的に追加
            if err_type not in self.detected_error_types:
                self.detected_error_types.add(err_type)
                current_values = sorted(list(self.detected_error_types - {"すべて"}))
                self.filter_type_combobox.config(values=["すべて"] + current_values)

            # フィルター条件と一致する場合のみTreeviewへ挿入
            if self._match_filter(error_item):
                self._insert_to_tree(error_item)

    def _match_filter(self, item):
        """現在のフィルター条件（エラー区分および検索キーワード）と合致しているか判定"""
        selected_type = self.filter_type_var.get()
        keyword = self.filter_keyword_var.get().strip().lower()

        # エラー区分による絞り込み
        if selected_type != "すべて" and item["type"] != selected_type:
            return False

        # キーワード検索による絞り込み (全列部分一致)
        if keyword:
            match_found = (
                keyword in item["basename"].lower() or
                keyword in item["sheet"].lower() or
                keyword in item["cell"].lower() or
                keyword in item["type"].lower() or
                keyword in item["detail"].lower()
            )
            if not match_found:
                return False

        return True

    def _insert_to_tree(self, item):
        """Treeviewへ行データを挿入し、偶数・奇数のストライプ模様を割り当てる"""
        current_count = len(self.tree.get_children())
        tag_stripe = "even" if current_count % 2 == 0 else "odd"
        self.tree.insert("", tk.END, values=(
            item["basename"],
            item["sheet"],
            item["cell"],
            item["type"],
            item["detail"]
        ), tags=(item["filepath"], tag_stripe))

    def _on_filter_changed(self, event=None):
        """フィルターが変更された際にTreeviewをクリアして再描画する"""
        for item in self.tree.get_children():
            self.tree.delete(item)

        for item in self.all_error_data:
            if self._match_filter(item):
                self._insert_to_tree(item)

    def _clear_filter(self):
        """すべてのフィルターをリセットして全件表示に戻す"""
        self.filter_type_var.set("すべて")
        self.filter_keyword_var.set("")
        self._on_filter_changed()

    def _finish_checker(self, errors_found, message):
        self.is_processing = False
        self.progress_bar["value"] = 100 if not self.cancel_requested else 0
        self.status_label.config(text=message)
        self.start_btn.config(state=tk.NORMAL, bg=self.primary_color)
        self.stop_btn.config(state=tk.DISABLED, bg="#9ca3af")
        self.settings_btn.config(state=tk.NORMAL)
        
        if self.cancel_requested:
            messagebox.showinfo("中止", "処理を中断しました。")
        elif errors_found > 0:
            messagebox.showwarning("チェック完了", f"入力ミスが {errors_found} 件検出されました。リストを確認してください。")
        else:
            messagebox.showinfo("チェック完了", "入力ミスは検出されませんでした。")

    def _show_detail_popup(self, event):
        selected_item = self.tree.selection()
        if not selected_item:
            return
        
        values = self.tree.item(selected_item[0], "values")
        if len(values) < 5:
            return
            
        popup = tk.Toplevel(self.root)
        popup.title("エラー詳細情報")
        popup.geometry("550x350")
        popup.transient(self.root)
        popup.grab_set()

        frm = ttk.Frame(popup, padding=15)
        frm.pack(fill=tk.BOTH, expand=True)

        labels = [
            ("ファイル名:", values[0]),
            ("シート名:", values[1]),
            ("セル位置:", values[2]),
            ("エラー種別:", values[3])
        ]

        for i, (lbl, val) in enumerate(labels):
            tk.Label(frm, text=lbl, font=("Yu Gothic UI", 10, "bold"), anchor="w").grid(row=i, column=0, sticky="nw", pady=5)
            tk.Label(frm, text=val, font=("Yu Gothic UI", 10), justify="left", anchor="w", wraplength=400).grid(row=i, column=1, sticky="nw", pady=5, padx=10)

        tk.Label(frm, text="詳細説明:", font=("Yu Gothic UI", 10, "bold"), anchor="w").grid(row=4, column=0, sticky="nw", pady=5)
        
        detail_txt = tk.Text(frm, font=("Yu Gothic UI", 10), height=6, wrap=tk.WORD, bg="#f9fafb", relief=tk.SOLID, bd=1)
        detail_txt.insert(tk.END, values[4])
        detail_txt.config(state=tk.DISABLED)
        detail_txt.grid(row=4, column=1, sticky="nsew", pady=5, padx=10)

        frm.rowconfigure(4, weight=1)
        frm.columnconfigure(1, weight=1)

        close_btn = ttk.Button(frm, text="閉じる", command=popup.destroy)
        close_btn.grid(row=5, column=0, columnspan=2, pady=(15, 0))

    # ==========================================
    # ヘルパーメソッド
    # ==========================================
    def _should_check_sheet(self, sheet_name):
        exclude_keywords = ["設定", "コード", "マスタ", "master", "list", "リスト", "summary", "集計"]
        name_lower = sheet_name.lower()
        return not any(k in name_lower for k in exclude_keywords)

    def _get_clean_value(self, val):
        if val is None:
            return ""
        if isinstance(val, float):
            if val.is_integer():
                return str(int(val))
            return str(val)
        return str(val).strip()

    # ==========================================
    # エクセル解析エンジン (絶対保存・上書きしない)
    # ==========================================
    def _check_excel_file(self, filepath, start_row_val, end_row_val, cols_info):
        errors = []
        ext = os.path.splitext(filepath)[1].lower()

        try:
            if ext in (".xlsx", ".xlsm"):
                errors.extend(self._parse_xlsx_xlsm(filepath, start_row_val, end_row_val, cols_info))
        except Exception as e:
            errors.append({
                "sheet": "ファイル全体",
                "cell": "N/A",
                "type": "読込失敗",
                "detail": f"Excelファイルを開くことができませんでした。破損またはパスワード保護の可能性があります。 (エラー: {str(e)})"
            })
        return errors

    def _parse_xlsx_xlsm(self, filepath, start_row_val, end_row_val, cols_info):
        errors = []
        # 日報ロジックチェックは data_only=True で実施
        wb = openpyxl.load_workbook(filepath, read_only=True, data_only=True)
        for sheet_name in wb.sheetnames:
            sheet = wb[sheet_name]
            
            if sheet.max_row is None or sheet.max_row == 0:
                continue
                
            if not self._should_check_sheet(sheet_name):
                continue

            # 1. 氏名・日付等のスマートスキャンチェック
            if self.cfg_check_smart_scan.get():
                self._scan_sheet_data_xlsx(sheet, sheet_name, errors)

            # 2. & 3. 管理No・コードの未入力整合性検証
            if self.cfg_check_mng_no.get() or self.cfg_check_code_integrity.get():
                self._check_code_integrity_xlsx(sheet, sheet_name, errors, start_row_val, end_row_val, cols_info)

            # 5. 時間単位チェック (倍数可変)
            if self.cfg_check_time_multiples.get():
                self._check_time_multiples_xlsx(sheet, sheet_name, errors)

            # 4. 数値フォーマットチェック (空白か数値以外の検証)
            if self.cfg_check_numeric_format.get():
                self._check_numeric_format_xlsx(sheet, sheet_name, errors, start_row_val, end_row_val, cols_info)
            
            # 6. 基準テンプレート比較
            if self.cfg_check_template_compare.get() and sheet_name in self.template_cache:
                self._compare_with_template_xlsx(sheet_name, errors, start_row_val, end_row_val, cols_info, filepath)
            
        wb.close()
        return errors

    # ------------------------------------------
    # 空白か数値以外のフォーマット検証ロジック (.xlsx 用)
    # ------------------------------------------
    def _check_numeric_format_xlsx(self, sheet, sheet_name, errors, start_row_val, end_row_val, cols_info):
        start_row = start_row_val
        end_row_limit = end_row_val if end_row_val is not None else sheet.max_row
        
        mng_idx = cols_info['mng']
        kat_idx = cols_info['kat']
        sag_idx = cols_info['sag']
        
        # M(12) 〜 AQ(42)  ※0-indexed
        m_to_aq_indices = list(range(12, 43))
        
        # チェック対象の列インデックス of セット
        check_col_indices = set([mng_idx, kat_idx, sag_idx] + m_to_aq_indices)
        check_col_indices = {c for c in check_col_indices if c >= 0}
        
        if not check_col_indices:
            return
            
        max_col_idx = max(check_col_indices)
        
        for r_idx, row in enumerate(sheet.iter_rows(min_row=start_row, max_row=end_row_limit, max_col=max_col_idx+1, values_only=True), start=start_row):
            for c_idx in check_col_indices:
                if c_idx < len(row):
                    val = row[c_idx]
                    val_str = self._get_clean_value(val)
                    if val_str == "" or val_str.lower() == "none":
                        continue
                    
                    try:
                        float(val_str)
                    except ValueError:
                        col_letter = openpyxl.utils.get_column_letter(c_idx + 1)
                        cell_pos = f"{col_letter}{r_idx}"
                        errors.append({
                            "sheet": sheet_name,
                            "cell": cell_pos,
                            "type": "数値フォーマットチェック",
                            "detail": f"空白か数値であるべきセルに、それ以外の値（'{val_str}'）が入力されています。"
                        })

    # ------------------------------------------
    # テンプレート書き換え検証ロジック (除外セル判定)
    # ------------------------------------------
    def _is_input_cell(self, r, c, start_row_val, end_row_val, cols_info):
        """セル位置 (r, c) [0-indexed] がユーザーの入力可能エリア（除外対象）か判定する"""
        start_row = start_row_val - 1
        end_row = end_row_val if end_row_val is not None else 100000

        # 指定されたテンプレート厳格チェック対象範囲は、入力セル（除外対象）から完全に除外する（必ず False を返す）
        # 1. G2からAS6: 2 <= 行 <= 6, G(7列) <= 列 <= AS(45列) (0-indexed: 1 <= r <= 5, 6 <= c <= 44)
        if (1 <= r <= 5) and (6 <= c <= 44):
            return False
            
        # 2. B7からAS7: 行 == 7, B(2列) <= 列 <= AS(45列) (0-indexed: r == 6, 1 <= c <= 44)
        if (r == 6) and (1 <= c <= 44):
            return False
            
        # 3. B4からL6: 4 <= 行 <= 6, B(2列) <= 列 <= L(12列) (0-indexed: 3 <= r <= 5, 1 <= c <= 11)
        if (3 <= r <= 5) and (1 <= c <= 11):
            return False
            
        # 4. J8からJ158: 8 <= 行 <= 158, 列 == J(10列) (0-indexed: 7 <= r <= 157, c == 9)
        if (7 <= r <= 157) and (c == 9):
            return False
            
        # 5. L8からL158: 8 <= 行 <= 158, 列 == L(12列) (0-indexed: 7 <= r <= 157, c == 11)
        if (7 <= r <= 157) and (c == 11):
            return False
            
        # 6. AR8からAR158: 8 <= 行 <= 158, 列 == AR(44列) (0-indexed: 7 <= r <= 157, c == 43)
        if (7 <= r <= 157) and (c == 43):
            return False
            
        # 7. A8からA158: 8 <= 行 <= 158, 列 == A(1列) (0-indexed: 7 <= r <= 157, c == 0)
        if (7 <= r <= 157) and (c == 0):
            return False

        # M6〜AR7（行インデックス 5, 6 かつ 列インデックス 12〜43）
        if r in (5, 6) and (12 <= c <= 43):
            return True
            
        # ユーザー指定の入力＆走査可能セルの除外範囲（8行目(7)〜158行目(157)）
        if start_row <= r < end_row:
            # B8からI158 (列インデックス B(1)〜I(8))
            if 1 <= c <= 8:
                return True
            # K8からK158 (列インデックス K(10))
            if c == 10:
                return True
            # M8からAQ158 (列インデックス M(12)〜AQ(42))
            if 12 <= c <= 42:
                return True
            # AS8からAS158 (列インデックス AS(44)) を除外
            if c == 44:
                return True
                    
        # 氏名・日付等の周辺（B3〜N3周辺など、日報ヘッダー部の入力欄を自動除外）
        if r in (2, 3) and (1 <= c <= 6):
            return True
            
        return False

    def _compare_with_template_xlsx(self, sheet_name, errors, start_row_val, end_row_val, cols_info, filepath):
        temp_data = self.template_cache[sheet_name]
        
        # 突き合わせ側は data_only=False で数式（=LOOKUP(...)等）を取得する
        wb_formula = openpyxl.load_workbook(filepath, read_only=True, data_only=False)
        if sheet_name not in wb_formula.sheetnames:
            wb_formula.close()
            return
        sheet_formula = wb_formula[sheet_name]
        rows_formula = list(sheet_formula.iter_rows(values_only=True))
        wb_formula.close()
        
        nrows = len(rows_formula)
        
        for r_idx in range(min(nrows, len(temp_data))):
            row_target = rows_formula[r_idx]
            row_temp = temp_data[r_idx]
            for c_idx in range(min(len(row_target), len(row_temp))):
                if self._is_input_cell(r_idx, c_idx, start_row_val, end_row_val, cols_info):
                    continue
                    
                val_target = self._get_clean_value(row_target[c_idx])
                val_temp = row_temp[c_idx]
                
                # 数式（=LOOKUP(...)）または固定値を比較
                if val_target != val_temp:
                    cell_pos = f"{openpyxl.utils.get_column_letter(c_idx + 1)}{r_idx + 1}"
                    errors.append({
                        "sheet": sheet_name,
                        "cell": cell_pos,
                        "type": "基準テンプレート不整合チェック",
                        "detail": f"入力セル以外の固定文字または数式が変更されています。[基準数式: '{val_temp}' / 対象数式: '{val_target}']"
                    })

    # ------------------------------------------
    # M6〜AR6, M7〜AR7 時間チェック用ロジック (.xlsx 用)
    # ------------------------------------------
    def _check_time_multiples_xlsx(self, sheet, sheet_name, errors):
        # ユーザー設定の倍数をパース
        try:
            multiple = int(self.cfg_time_multiple_val.get().strip())
            if multiple <= 0:
                multiple = 60
        except ValueError:
            multiple = 60

        # 6行目と7行目をチェック
        for r_idx in [6, 7]:
            # M(13)からAR(44)のセル値を取得。openpyxlは1始まり。
            row_cells = list(sheet.iter_rows(min_row=r_idx, max_row=r_idx, min_col=13, max_col=44, values_only=True))
            if not row_cells or len(row_cells) == 0:
                continue
            
            row_values = row_cells[0]
            for idx, val in enumerate(row_values):
                col_idx = 13 + idx
                cell_pos = f"{openpyxl.utils.get_column_letter(col_idx)}{r_idx}"
                
                val_str = self._get_clean_value(val)
                if val_str == "" or val_str.lower() == "none":
                    continue
                
                try:
                    num_val = float(val_str)
                    if not num_val.is_integer():
                        errors.append({
                            "sheet": sheet_name,
                            "cell": cell_pos,
                            "type": "勤務時間単位チェック",
                            "detail": f"{r_idx}行目のセル値（{val_str}）は整数（0または{multiple}の倍数）ではありません。"
                        })
                        continue
                    
                    num_int = int(num_val)
                    if num_int != 0 and num_int % multiple != 0:
                        errors.append({
                            "sheet": sheet_name,
                            "cell": cell_pos,
                            "type": "勤務時間単位チェック",
                            "detail": f"{r_idx}行目のセル値（{num_int}）は0または{multiple}の倍数ではありません。"
                        })
                except ValueError:
                    errors.append({
                        "sheet": sheet_name,
                        "cell": cell_pos,
                        "type": "勤務時間単位チェック",
                        "detail": f"{r_idx}行目のセルに数値以外の値（'{val_str}'）が入力されています。"
                    })

    # ------------------------------------------
    # 管理Noと型枠/作業コードの未入力整合性検証 (.xlsx 用)
    # ------------------------------------------
    def _check_code_integrity_xlsx(self, sheet, sheet_name, errors, start_row_val, end_row_val, cols_info):
        # 必要な範囲のみをロードするよう変更（高速化と範囲の正確化）
        start_row = start_row_val
        end_row_limit = end_row_val if end_row_val is not None else sheet.max_row
        
        mng_idx = cols_info['mng']
        kat_idx = cols_info['kat']
        sag_idx = cols_info['sag']
        tot_idx = cols_info['tot']

        # 最大インデックスの計算
        # C〜H列(インデックス2〜7)もチェックするため、max_col_idxが7以上になるように保証する
        max_col_idx = max(mng_idx, kat_idx, sag_idx, tot_idx, 7)
        if max_col_idx < 0:
            return # 列が指定されていない場合はスキップ

        for r_idx, row in enumerate(sheet.iter_rows(min_row=start_row, max_row=end_row_limit, max_col=max_col_idx+1, values_only=True), start=start_row):
            mng_val = ""
            if 0 <= mng_idx < len(row):
                mng_val = self._get_clean_value(row[mng_idx])
                
            is_valid_mng = False
            if mng_val and mng_val != "None" and mng_val != "":
                if mng_val.isdigit():
                    val_num = int(mng_val)
                    if not (40000 <= val_num <= 50000):
                        is_valid_mng = True
                elif len(mng_val) >= 4:
                    is_valid_mng = True
                            
            # 合計値（実績）の算出（固定された指定列の値のみを見る）
            total_num = 0.0
            if 0 <= tot_idx < len(row):
                total_val = self._get_clean_value(row[tot_idx])
                try:
                    total_num = float(total_val)
                except ValueError:
                    pass
                        
            kat_val = ""
            if 0 <= kat_idx < len(row):
                kat_val = self._get_clean_value(row[kat_idx])
                    
            text_sag_val = ""
            if 0 <= sag_idx < len(row):
                text_sag_val = self._get_clean_value(row[sag_idx])
                    
            is_kat_empty = (kat_val == "" or kat_val == "None" or kat_val == "0")
            is_sag_empty = (text_sag_val == "" or text_sag_val == "None" or text_sag_val == "0")

            # C〜H列 (インデックス 2〜7) の空白以外チェック
            c_to_h_has_value = False
            for c_idx in range(2, 8):
                if c_idx < len(row):
                    val_str = self._get_clean_value(row[c_idx])
                    if val_str != "" and val_str.lower() != "none":
                        c_to_h_has_value = True
                        break

            # 新規要件: C,D,E,F,G,Hが空白以外で、B,I,Kが空白の場合にエラー (コード整合性チェックの一部)
            if self.cfg_check_code_integrity.get() and c_to_h_has_value and not is_valid_mng and is_kat_empty and is_sag_empty:
                col_letter = openpyxl.utils.get_column_letter(mng_idx + 1) if mng_idx >= 0 else "?"
                cell_pos_str = f"{col_letter}{r_idx}"
                errors.append({
                    "sheet": sheet_name,
                    "cell": cell_pos_str,
                    "type": "コード未入力チェック",
                    "detail": f"{r_idx}行目のC〜H列に記述がありますが、管理No(B)、型枠コード(I)、作業コード(K)がすべて未入力です。"
                })
                continue # 既存のエラーと重複しないようスキップ

            # 2. 管理No未入力チェックが有効な場合
            if self.cfg_check_mng_no.get() and not is_valid_mng and total_num > 0:
                col_letter = openpyxl.utils.get_column_letter(mng_idx + 1) if mng_idx >= 0 else "?"
                cell_pos_str = f"{col_letter}{r_idx}"
                errors.append({
                    "sheet": sheet_name,
                    "cell": cell_pos_str,
                    "type": "管理No未入力チェック",
                    "detail": f"作業実績（合計 {total_num}）が入力されていますが、管理Noが未入力です。"
                })

            # 3. コード未入力チェックが有効な場合
            elif self.cfg_check_code_integrity.get() and is_valid_mng:
                # 型枠コード、あるいは作業コードの「両方とも」が未入力の場合にエラーとする
                if is_kat_empty and is_sag_empty:
                    col_letter = openpyxl.utils.get_column_letter(mng_idx + 1) if mng_idx >= 0 else "?"
                    cell_pos_str = f"{col_letter}{r_idx}"
                    errors.append({
                        "sheet": sheet_name,
                        "cell": cell_pos_str,
                        "type": "コード未入力チェック",
                        "detail": f"管理No '{mng_val}' が指定されていますが、型枠コードと作業コードの両方が未入力です。"
                    })

    # ------------------------------------------
    # 汎用スマートスキャンルールエンジン
    # ------------------------------------------
    def _scan_sheet_data_xlsx(self, sheet, sheet_name, errors):
        for r_idx, row in enumerate(sheet.iter_rows(max_row=200, max_col=50, values_only=True), start=1):
            for c_idx, cell_val in enumerate(row, start=1):
                if cell_val is None or not isinstance(cell_val, str):
                    continue
                
                clean_val = cell_val.replace(" ", "").replace("　", "")
                
                if any(k in clean_val for k in ["氏名", "担当者", "名前", "記述者", "作成者"]) and "印" not in clean_val:
                    self._validate_neighbor_xlsx(sheet, sheet_name, r_idx, c_idx, "氏名", errors)

                if any(k in clean_val for k in ["日付", "年月日", "作成日", "報告日"]):
                    self._validate_neighbor_xlsx(sheet, sheet_name, r_idx, c_idx, "日付", errors)

                if any(k == clean_val for k in ["時間", "工数", "作業時間", "h", "H"]):
                    self._validate_table_column_xlsx(sheet, sheet_name, r_idx, c_idx, errors)

    def _validate_neighbor_xlsx(self, sheet, sheet_name, r, c, item_type, errors):
        candidates = []
        candidates.append((r, c + 1, "右隣"))
        candidates.append((r + 1, c, "下"))

        validated = False
        for target_r, target_c, direction in candidates:
            val = sheet.cell(row=target_r, column=target_c).value
            if val is not None and str(val).strip() != "":
                val_str = str(val).strip()
                cell_pos_str = f"{openpyxl.utils.get_column_letter(target_c)}{target_r}"
                
                if item_type == "日付":
                    if isinstance(val, (datetime, tk.Variable)):
                        pass
                    else:
                        if not self._is_valid_date_string(val_str):
                            errors.append({
                                "sheet": sheet_name,
                                "cell": cell_pos_str,
                                "type": "氏名・日付等スマートスキャン",
                                "detail": f"「{item_type}」セルの{direction}に有効な日付が入力されていません。入力値: '{val_str}'"
                            })
                validated = True
                break

        if not validated:
            col_letter = openpyxl.utils.get_column_letter(c)
            errors.append({
                "sheet": sheet_name,
                "cell": f"{col_letter}{r}",
                "type": "氏名・日付等スマートスキャン",
                "detail": f"「{item_type}」セルの周辺（右隣または下）に有効なデータが入力されていません。"
            })

    def _validate_table_column_xlsx(self, sheet, sheet_name, header_r, header_c, errors):
        consecutive_empty = 0
        r = header_r + 1
        
        while consecutive_empty < 3:
            val = sheet.cell(row=r, column=header_c).value
            cell_pos_str = f"{openpyxl.utils.get_column_letter(header_c)}{r}"
            
            other_data_filled = False
            for col_idx in range(1, 20):
                if col_idx != header_c:
                    other_val = sheet.cell(row=r, column=col_idx).value
                    if other_val is not None and str(other_val).strip() != "":
                        other_data_filled = True
                        break
            
            if val is None or str(val).strip() == "":
                if other_data_filled:
                    errors.append({
                        "sheet": sheet_name,
                        "cell": cell_pos_str,
                        "type": "氏名・日付等スマートスキャン",
                        "detail": "業務記述行ですが、作業時間(工数)が空欄になっています。"
                    })
                consecutive_empty += 1
            else:
                consecutive_empty = 0
                try:
                    num_val = float(val)
                    if num_val < 0:
                        errors.append({
                            "sheet": sheet_name,
                            "cell": cell_pos_str,
                            "type": "氏名・日付等スマートスキャン",
                            "detail": f"作業時間に負の値が入力されています。入力値: {num_val}"
                        })
                    elif num_val > 24:
                        errors.append({
                            "sheet": sheet_name,
                            "cell": cell_pos_str,
                            "type": "氏名・日付等スマートスキャン",
                            "detail": f"1日の作業工数として過大な時間(24h超)が入力されています。入力値: {num_val}"
                        })
                except ValueError:
                    errors.append({
                        "sheet": sheet_name,
                        "cell": cell_pos_str,
                        "type": "氏名・日付等スマートスキャン",
                        "detail": f"時間入力欄に数字以外の値が入力されています。入力値: '{val}'"
                    })
            r += 1

    def _is_valid_date_string(self, text):
        patterns = [
            r"^\d{4}[-/]\d{1,2}[-/]\d{1,2}$",
            r"^\d{1,2}[-/]\d{1,2}$",
            r"^\d{4}年\d{1,2}月\d{1,2}日$",
            r"^\d{1,2}月\d{1,2}日$"
        ]
        text_clean = text.strip()
        for p in patterns:
            if re.match(p, text_clean):
                return True
        return False

if __name__ == "__main__":
    root = tk.Tk()
    app = DailyReportCheckerApp(root)
    root.mainloop()